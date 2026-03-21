"""
Biological Report Extractor
============================
Uses LlamaParse (for PDF parsing) + Google ADK (for agent/LLM extraction)
to extract structured fields from biological/environmental PDF reports
and write them to an Excel file.

Requirements:
    pip install llama-parse llama-index-core google-adk google-genai openpyxl python-dotenv

Setup:
    Create a .env file in this folder with:
        LLAMA_CLOUD_API_KEY=your_llamaparse_key
        GOOGLE_API_KEY=your_gemini_key
"""

import os
import json
import asyncio
from pathlib import Path
from dotenv import load_dotenv

# Pypdf
from pypdf import PdfReader

# Google ADK
from google.adk.agents import LlmAgent
from google.adk.runners import Runner
from google.adk.sessions import InMemorySessionService
from google.genai import types as genai_types

# Excel output
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

load_dotenv()

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────

PDF_PATH = r"C:\pdf_agent\dsd_biological-technical-report_1.pdf"   # <-- change to your PDF path
OUTPUT_EXCEL = "bio_report_output.xlsx"
GEMINI_MODEL = "gemini-2.0-flash"

# Summary fields only — species handled separately
FIELDS = {
    "Title":            "The full title of the report",
    "Author":           "The person or agency who wrote/prepared the report",
    "Client":           "The person or agency this report was produced for (the client)",
    "Purpose":          "A brief description of the project goals. Maximum two sentences.",
    "Publication_Date": "The date the report was published or completed",
    "Location":         "The specific location being studied or addressed in the report",
}


# ─────────────────────────────────────────────
# PROMPTS
# ─────────────────────────────────────────────

EXTRACTION_PROMPT = """
You are a specialist in extracting structured data from biological and environmental reports.

Here is the full content of a biological technical report:

<report>
{report_text}
</report>

Extract the following fields from the report. For each field, provide your answer based ONLY
on what is explicitly stated in the document. If a field cannot be found, use "NOT FOUND".

Fields to extract:
{field_descriptions}

Respond ONLY with a valid JSON object with exactly these keys:
{field_keys}

Rules:
- Never use HTML entities in any field. Always decode them to plain text characters (e.g. & not &#x26;).
- For all fields, output plain text only. No HTML encoding, no special character escaping.
- Publication_Date: use the format shown in the document (e.g. "September 2024")
- Purpose: maximum two sentences, in your own words summarising the project
- Location: Follow this priority order strictly and use the first option you can find:
  (1) Street address including city and state e.g. "1234 Main St, San Diego, CA"
  (2) If no street address exists, use the project/site name plus city and state e.g. "El Camino Memorial Park, San Diego, CA"
  (3) If neither exists, use city and state only e.g. "San Diego, CA"
  Never combine multiple formats or list multiple locations.
- Do not add any explanation outside the JSON object
"""

SPECIES_PROMPT = """
You are extracting a species list from a biological report.

<report>
{report_text}
</report>

Your ONLY task is to extract species that were DIRECTLY OBSERVED or DETECTED on site during surveys.

Focus your search on: {target_sections}
If this specifies particular tables or appendices, extract ONLY from those sections.
If it says "the entire document", scan everything.

INCLUDE species that were:
- Observed visually during field surveys
- Detected by sound or other evidence during surveys
- Found as physical evidence on site (nests, tracks, scat)
- Recorded as present in the study area

EXCLUDE species that were:
- Listed as having "potential to occur" or "may occur"
- Listed as "not observed" or "not detected"
- Mentioned only in regional context or background literature
- Listed in tables analysing species that were considered but not found
- Described as occurring near the site but not on it

Before writing your answer, scan the entire document twice:
- First pass: collect all species from tables and species lists
- Second pass: collect any additional species mentioned in the narrative text

For each species record:
- The species name using format "Common Name (Scientific Name)" if both are available, or just "Common Name" if no scientific name is available
- The section of the report where it was found e.g. "Section 3.6 Flora", "Appendix B Fauna List", "Table 2"

Return ONLY a JSON object in this exact format:
{{
  "Species_Names": [
    {{"name": "Coast Live Oak (Quercus agrifolia)", "source": "Section 3.5.2"}},
    {{"name": "Coastal California Gnatcatcher (Polioptila californica)", "source": "Section 3.9.2"}}
  ]
}}

Do not summarise. Do not skip rows. Every observed species must appear in your output.
"""


# ─────────────────────────────────────────────
# STEP 1: Parse PDF with pypdf
# ─────────────────────────────────────────────

def parse_pdf(pdf_path: str) -> str:
    print(f"[1/3] Parsing PDF: {pdf_path}")
    reader = PdfReader(pdf_path)
    pages = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            pages.append(text)
    full_text = "\n\n".join(pages)
    print(f"    Parsed {len(pages)} pages, {len(full_text):,} characters")
    return full_text

# ─────────────────────────────────────────────
# STEP 2: Extract fields with Google ADK Agent
# ─────────────────────────────────────────────

async def extract_fields_with_adk(report_text: str, target_sections: str = "") -> dict:
    """Use Google ADK LlmAgent to extract structured fields from report text."""
    print("[2/3] Running ADK agent extraction...")

    

    # Build the field descriptions string
    field_descriptions = "\n".join(
        f'- {key}: {desc}' for key, desc in FIELDS.items()
    )
    field_keys = json.dumps(list(FIELDS.keys()), indent=2)

    # Truncate report text if very long
    max_chars = 800_000
    if len(report_text) > max_chars:
        print(f"    Report truncated from {len(report_text):,} to {max_chars:,} chars")
        report_text = report_text[:max_chars]

    prompt = EXTRACTION_PROMPT.format(
        report_text=report_text,
        field_descriptions=field_descriptions,
        field_keys=field_keys,
    )

    # Set up ADK agent
    agent = LlmAgent(
        name="bio_extractor",
        model=GEMINI_MODEL,
        description="Extracts structured fields from biological reports",
        instruction="You are a precise data extractor. Always return valid JSON only.",
        generate_content_config={"max_output_tokens": 8000},
    )

    session_service = InMemorySessionService()
    await session_service.create_session(
        app_name="bio_extractor",
        user_id="user_1",
        session_id="session_1",
    )

    runner = Runner(
        agent=agent,
        app_name="bio_extractor",
        session_service=session_service,
    )

    # ── First pass: summary fields ──
    message = genai_types.Content(
        role="user",
        parts=[genai_types.Part(text=prompt)]
    )

    response_text = ""
    async for event in runner.run_async(
        user_id="user_1",
        session_id="session_1",
        new_message=message,
    ):
        if event.is_final_response():
            if event.content and event.content.parts:
                response_text = event.content.parts[0].text
            break

    # Parse JSON response — find the first { ... } object in the output
    raw = response_text.strip()
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
    raw = raw.strip()

    start = raw.find("{")
    clean = raw[start:] if start != -1 else raw

    # If JSON is incomplete, try to close it
    if not clean.endswith("}"):
        last_comma = clean.rfind('",')
        if last_comma > 0:
            clean = clean[:last_comma+1]
        clean = clean + '\n"_truncated": "true"\n}'

    try:
        extracted = json.loads(clean)
        print(f"    Extracted {len(extracted)} fields successfully")
    except json.JSONDecodeError as e:
        print(f"    WARNING: Could not parse JSON response: {e}")
        print(f"    Raw response:\n{response_text[:500]}")
        extracted = {k: "EXTRACTION FAILED - check raw output" for k in FIELDS}

    # ── Second pass: species only ──
    print("    Running second pass for species extraction...")
    await session_service.create_session(
        app_name="bio_extractor",
        user_id="user_1",
        session_id="session_2",
    )
    species_prompt = SPECIES_PROMPT.format(
        report_text=report_text,
        target_sections=target_sections if target_sections else "the entire document",
    )

    message2 = genai_types.Content(
        role="user",
        parts=[genai_types.Part(text=species_prompt)]
    )

    response_text2 = ""
    async for event in runner.run_async(
        user_id="user_1",
        session_id="session_2",
        new_message=message2,
    ):
        if event.is_final_response():
            if event.content and event.content.parts:
                response_text2 = event.content.parts[0].text
            break

    clean2 = response_text2.strip()
    if clean2.startswith("```"):
        clean2 = clean2.split("```")[1]
        if clean2.startswith("json"):
            clean2 = clean2[4:]
    clean2 = clean2.strip()

    try:
        species_data = json.loads(clean2)
        extracted["Species_Names"] = species_data.get("Species_Names", "NOT FOUND")
        print(f"    Species pass complete")
    except json.JSONDecodeError:
        print(f"    WARNING: Species pass failed")
        extracted["Species_Names"] = "NOT FOUND"

    return extracted


# ─────────────────────────────────────────────
# STEP 3: Write results to Excel
# ─────────────────────────────────────────────

def write_to_excel(extracted: dict, output_path: str, source_pdf: str):
    """Write extracted fields to a formatted Excel file."""
    print(f"[3/3] Writing results to: {output_path}")

    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="2E4057")
    field_font  = Font(name="Arial", bold=True, size=10)
    value_font  = Font(name="Arial", size=10)
    title_font  = Font(name="Arial", bold=True, size=14, color="2E4057")
    sub_font    = Font(name="Arial", italic=True, size=9, color="888888")
    wrap        = Alignment(wrap_text=True, vertical="top")
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
    )

    # ── Sheet 1: Report Summary ──────────────────────────────
    ws = wb.active
    ws.title = "Report Summary"

    ws.merge_cells("A1:B1")
    ws["A1"] = "Biological Report — Extracted Data"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:B2")
    ws["A2"] = f"Source: {Path(source_pdf).name}"
    ws["A2"].font = sub_font
    ws.row_dimensions[2].height = 16

    ws.row_dimensions[4].height = 22
    for col, label in enumerate(["Field", "Extracted Value"], start=1):
        cell = ws.cell(row=4, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    row = 5
    for field_key in FIELDS:
        value = extracted.get(field_key, "NOT FOUND")

        ws.cell(row=row, column=1, value=field_key.replace("_", " ")).font = field_font
        ws.cell(row=row, column=1).alignment = wrap
        ws.cell(row=row, column=1).border = thin_border

        ws.cell(row=row, column=2, value=value).font = value_font
        ws.cell(row=row, column=2).alignment = wrap
        ws.cell(row=row, column=2).border = thin_border

        ws.row_dimensions[row].height = max(20, min(80, len(str(value)) // 3))
        row += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 80

    # ── Sheet 2: Species List ──────────────────────────
    ws2 = wb.create_sheet("Species List")

    ws2.merge_cells("A1:C1")
    ws2["A1"] = "Species Observed / Detected in Report"
    ws2["A1"].font = title_font
    ws2["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[1].height = 30

    ws2.row_dimensions[3].height = 22
    for col, label in enumerate(["#", "Species Name", "Source"], start=1):
        cell = ws2.cell(row=3, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Parse species list
    raw_species = extracted.get("Species_Names", "NOT FOUND")
    if isinstance(raw_species, list) and len(raw_species) > 0:
        species_list = [(s.get("name", ""), s.get("source", "")) for s in raw_species]
    elif isinstance(raw_species, str) and raw_species != "NOT FOUND":
        species_list = [(s.strip(), "") for s in raw_species.split("|") if s.strip()]
    else:
        species_list = [("No species data extracted", "")]

    alt_fill = PatternFill("solid", start_color="F5F7FA")

    for i, (name, source) in enumerate(species_list, start=1):
        r = i + 3
        ws2.cell(row=r, column=1, value=i).font = value_font
        ws2.cell(row=r, column=1).alignment = Alignment(horizontal="center")
        ws2.cell(row=r, column=2, value=name).font = value_font
        ws2.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws2.cell(row=r, column=3, value=source).font = value_font
        ws2.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="top")
        if i % 2 == 0:
            for col in range(1, 4):
                ws2.cell(row=r, column=col).fill = alt_fill
        ws2.row_dimensions[r].height = 16

    ws2.column_dimensions["A"].width = 6
    ws2.column_dimensions["B"].width = 55
    ws2.column_dimensions["C"].width = 40

    note_row = len(species_list) + 5
    ws2.cell(row=note_row, column=1, value=f"Total species: {len(species_list)}").font = Font(
        name="Arial", italic=True, size=9, color="888888"
    )

    wb.save(output_path)
    print(f"    Saved: {output_path}")
    print(f"    Summary fields: {len(FIELDS)}")
    print(f"    Species extracted: {len(species_list)}")


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

async def main():
    print("=" * 55)
    print("  Biological Report Field Extractor")
    print("=" * 55)

   # 1. Parse PDF
    report_text = parse_pdf(PDF_PATH)

    # 2. Ask user for target sections
    print("\nWhich tables or sections contain the directly observed species?")
    print("e.g. 'Appendix A, Table 3' or press Enter to scan the whole document")
    target_sections = input("> ").strip()

    if target_sections:
        print(f"    Targeting: {target_sections}")
    else:
        print("    No sections specified — scanning full document")

    # 3. Extract with ADK agent
    extracted = await extract_fields_with_adk(report_text, target_sections)

    # Print preview
    print("\n-- Extraction Preview --")
    for key, val in extracted.items():
        if key == "Species_Names":
            count = len(val) if isinstance(val, list) else len(str(val).split("|"))
            print(f"  {key}: {count} species found")
        else:
            preview = str(val)[:120] + "..." if len(str(val)) > 120 else str(val)
            print(f"  {key}: {preview}")

    # 3. Write to Excel
    write_to_excel(extracted, OUTPUT_EXCEL, PDF_PATH)

    print("\nDone! Open", OUTPUT_EXCEL, "to review results.")


if __name__ == "__main__":
    asyncio.run(main())